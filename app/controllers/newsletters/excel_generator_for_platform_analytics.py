import io
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

HERE = Path(__file__).resolve().parent
APP_ROOT = HERE.parent.parent

TEMPLATE_PATH = APP_ROOT / "quixote-report-template.xlsx"


def dataframe_to_excel_stream(
        analytics_data,
        template_path: Path = TEMPLATE_PATH
) -> io.BytesIO:
    """Return a BytesIO containing the filled‑in template workbook."""
    df = pd.DataFrame(analytics_data)
    wb = openpyxl.load_workbook(template_path)
    ws = wb["raw-data-tracking"]  # sheet that holds Table1
    tbl = ws.tables["Table1"]  # pivot source table

    # ---- clear old data (everything below the header row) ----
    header_row = ws[tbl.ref.split(":")[0]].row
    ws.delete_rows(header_row + 1, ws.max_row - header_row)

    # ---- write dataframe rows ----
    for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), header_row + 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # ---- resize table range ----
    last_row = header_row + len(df)
    last_col_letter = openpyxl.utils.get_column_letter(df.shape[1])
    tbl.ref = f"A{header_row}:{last_col_letter}{last_row}"

    # ---- save to BytesIO ----
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
